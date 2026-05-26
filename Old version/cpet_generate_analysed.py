#!/usr/bin/env python3
"""
Generate an analysed CPET Excel workbook from a raw appended CPET file.

Expected raw workbook layout:
    5s AVERAGED DATA
    BP DATA
    BREATH x BREATH DATA

Usage:
    python cpet_generate_analysed.py FRIENDS_009_CPET_Files.XLS
    python cpet_generate_analysed.py FRIENDS_009_CPET_Files.xlsx -o output_ANALYSED.xlsx
"""

from __future__ import annotations

import argparse
import csv
import re
from datetime import datetime, time, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


SUMMARY_ROWS = 15
ROLLING_WINDOW_ROWS = 6  # 30 seconds in the 5-sec averaged file

# Final analysed column positions after inserting calculated columns.
COL_TIME = 1
COL_LOAD = 2
COL_HR = 3
COL_VE = 4
COL_VO2 = 6
COL_VO2_ROLLING = 7
COL_VCO2 = 8
COL_RER = 9
COL_VO2KG = 10
COL_VO2KG_ROLLING = 11
COL_O2PULSE = 12

# BP columns after inserting calculated columns globally.
COL_BP_SYS = 5
COL_BP_DIA = 6


class CPETAnalysisError(RuntimeError):
    """Raised when the raw workbook structure cannot be interpreted."""


def clean_text(value: Any) -> str:
    return str(value).strip().lower() if value is not None else ""


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def to_number(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None

    if is_number(value):
        return float(value)

    try:
        return float(str(value).strip())
    except Exception:
        return None


def parse_time_to_seconds(value: Any) -> Optional[int]:
    """
    Parse CPET time values such as:
        09:57
        9:57
        00:09:57
        Excel time cells
    """
    if value is None:
        return None

    if isinstance(value, time):
        return value.hour * 3600 + value.minute * 60 + value.second

    if isinstance(value, datetime):
        return value.hour * 3600 + value.minute * 60 + value.second

    if isinstance(value, timedelta):
        return int(round(value.total_seconds()))

    # Excel may store times as a fraction of a day.
    if is_number(value):
        numeric = float(value)
        if 0 <= numeric < 1:
            return int(round(numeric * 24 * 60 * 60))
        return None

    text = str(value).strip()

    match = re.fullmatch(r"(\d{1,2}):(\d{2})(?::(\d{2}))?", text)
    if not match:
        return None

    first = int(match.group(1))
    second = int(match.group(2))
    third = int(match.group(3)) if match.group(3) is not None else None

    if third is None:
        # CPET export usually uses mm:ss.
        return first * 60 + second

    # If hh:mm:ss is provided.
    return first * 3600 + second * 60 + third


def coerce_cell(value: str) -> Any:
    """Convert tab-delimited text values into numbers where possible."""
    if value is None:
        return None

    text = str(value).rstrip("\n\r")

    if text == "":
        return None

    # Preserve time strings such as ' 09:57'.
    if re.fullmatch(r"\s*\d{1,2}:\d{2}(?::\d{2})?\s*", text):
        return text

    try:
        num = float(text)
        if num.is_integer():
            return int(num)
        return num
    except Exception:
        return text


def load_raw_file(path: Path) -> Workbook:
    """Load .xlsx normally; load CPET .XLS export as tab-delimited text."""
    suffix = path.suffix.lower()

    if suffix in {".xlsx", ".xlsm"}:
        return load_workbook(path)

    raw = path.read_text(encoding="utf-8-sig", errors="replace")

    wb = Workbook()
    ws = wb.active
    ws.title = path.stem[:31]

    reader = csv.reader(raw.splitlines(), delimiter="\t")
    for row in reader:
        ws.append([coerce_cell(cell) for cell in row])

    return wb


def find_row(ws, label: str) -> int:
    target = clean_text(label)

    for row in range(1, ws.max_row + 1):
        if clean_text(ws.cell(row, 1).value) == target:
            return row

    raise CPETAnalysisError(f"Could not find block label: {label!r}")


def block_rows(ws, label: str, next_label: Optional[str] = None) -> Dict[str, int]:
    label_row = find_row(ws, label)
    header_row = label_row + 1
    unit_row = label_row + 2
    data_start = label_row + 3

    if next_label:
        next_label_row = find_row(ws, next_label)
        data_end = next_label_row - 1
    else:
        data_end = ws.max_row

    # Trim trailing blank rows.
    while data_end >= data_start:
        row_values = [
            ws.cell(data_end, c).value
            for c in range(1, min(ws.max_column, 14) + 1)
        ]

        if any(v is not None for v in row_values):
            break

        data_end -= 1

    return {
        "label": label_row,
        "header": header_row,
        "unit": unit_row,
        "start": data_start,
        "end": data_end,
    }


def remove_irregular_5s_rows(ws, start: int, end: int, time_col: int = COL_TIME) -> int:
    """
    Delete non-5-second timestamp rows from the 5s averaged data block.

    Example:
        09:55
        09:57  <- deleted
        10:00

    This is applied only to the 5s AVERAGED DATA block.
    """
    rows_to_delete: List[int] = []

    for row in range(start, end + 1):
        value = ws.cell(row, time_col).value
        seconds = parse_time_to_seconds(value)

        if seconds is None:
            continue

        if seconds % 5 != 0:
            rows_to_delete.append(row)

    for row in reversed(rows_to_delete):
        ws.delete_rows(row, 1)

    return len(rows_to_delete)


def find_irregular_5s_rows(ws, start: int, end: int, time_col: int = COL_TIME) -> List[Tuple[int, Any]]:
    """Check whether irregular 5s rows remain."""
    irregular: List[Tuple[int, Any]] = []

    for row in range(start, end + 1):
        value = ws.cell(row, time_col).value
        seconds = parse_time_to_seconds(value)

        if seconds is None:
            continue

        if seconds % 5 != 0:
            irregular.append((row, value))

    return irregular


def final_window_start_row(
    block_start: int,
    peak_row: int,
    window_rows: int = ROLLING_WINDOW_ROWS,
) -> int:
    """
    Final 30 sec window = exactly 6 rows including the peak row.

    Example:
        peak row = 149
        start row = 149 - 6 + 1 = 144
        formula = AVERAGE(row 144:row 149)
    """
    return max(block_start, peak_row - window_rows + 1)


def first_load_gt_zero_row(ws, start: int, end: int, load_col: int = COL_LOAD) -> Optional[int]:
    for row in range(start, end + 1):
        load = to_number(ws.cell(row, load_col).value)

        if load is not None and load > 0:
            return row

    return None


def last_zero_load_before_exercise(
    ws,
    start: int,
    end: int,
    load_col: int = COL_LOAD,
) -> Optional[int]:
    first_exercise_row = first_load_gt_zero_row(ws, start, end, load_col)

    if first_exercise_row:
        search_end = first_exercise_row - 1
    else:
        search_end = end

    candidates = []

    for row in range(start, search_end + 1):
        load = to_number(ws.cell(row, load_col).value)

        if load == 0:
            candidates.append(row)

    return candidates[-1] if candidates else None


def max_load_row(ws, start: int, end: int, load_col: int = COL_LOAD) -> int:
    """
    Return the LAST row where the maximum load occurs.

    This is important because peak load may be sustained for several rows.
    We want calculations to include data until the final occurrence of peak load,
    not stop at the first occurrence.
    """
    max_value = None
    max_row = None

    for row in range(start, end + 1):
        value = to_number(ws.cell(row, load_col).value)

        if value is None:
            continue

        if max_value is None or value > max_value:
            max_value = value
            max_row = row

        elif value == max_value:
            # Keep updating so we end at the LAST peak-load row.
            max_row = row

    if max_row is None:
        raise CPETAnalysisError(f"No numeric load values found in rows {start}:{end}")

    return max_row


def write_rolling_average_formulas(ws, start: int, end: int) -> None:
    """Write 30-sec rolling average formulas beside VO2 and VO2/kg."""
    for row in range(start, end + 1):
        vo2_start = max(start, row - ROLLING_WINDOW_ROWS + 1)
        vo2kg_start = max(start, row - ROLLING_WINDOW_ROWS + 1)

        ws.cell(row, COL_VO2_ROLLING).value = f"=AVERAGE(F{vo2_start}:F{row})"
        ws.cell(row, COL_VO2KG_ROLLING).value = f"=AVERAGE(J{vo2kg_start}:J{row})"

        ws.cell(row, COL_VO2_ROLLING).number_format = "0.0"
        ws.cell(row, COL_VO2KG_ROLLING).number_format = "0.0"


def apply_basic_style(ws) -> None:
    """Make the workbook visually close to the manually analysed file."""
    ws.column_dimensions["A"].width = 22

    for col in range(2, 15):
        ws.column_dimensions[get_column_letter(col)].width = 12

    bold = Font(bold=True)
    title_fill = PatternFill("solid", fgColor="D9EAF7")
    thin_gray = Side(style="thin", color="BFBFBF")
    border = Border(
        left=thin_gray,
        right=thin_gray,
        top=thin_gray,
        bottom=thin_gray,
    )

    # Summary block styling.
    for row in [1, 5]:
        ws.cell(row, 1).font = bold
        ws.cell(row, 1).fill = title_fill

    for row in range(1, 15):
        ws.cell(row, 1).alignment = Alignment(horizontal="left")
        ws.cell(row, 2).number_format = "0.0"

    # RER summary value as 2 decimal places.
    ws.cell(10, 2).number_format = "0.00"

    # Block headings and headers.
    for label in ["5s AVERAGED DATA", "BP DATA", "BREATH x BREATH DATA"]:
        try:
            r = find_row(ws, label)

            ws.cell(r, 1).font = bold
            ws.cell(r, 1).fill = title_fill

            for c in range(1, 15):
                ws.cell(r + 1, c).font = bold
                ws.cell(r + 1, c).border = border
                ws.cell(r + 2, c).font = Font(italic=True)
                ws.cell(r + 2, c).border = border

        except CPETAnalysisError:
            pass


def build_analysed_workbook(input_path: Path, output_path: Path) -> None:
    wb = load_raw_file(input_path)
    ws = wb.active

    ws.title = f"{input_path.stem}_ANALYSED"[:31]

    # Add 15 summary rows above the raw appended data.
    ws.insert_rows(1, SUMMARY_ROWS)

    # Insert calculated rolling-average columns globally.
    ws.insert_cols(COL_VO2_ROLLING, 1)       # G, beside VO2 in F
    ws.insert_cols(COL_VO2KG_ROLLING, 1)    # K, beside VO2/kg now in J

    # Detect 5s block after row/column insertion.
    five = block_rows(ws, "5s AVERAGED DATA", "BP DATA")

    # Delete irregular 5s rows such as 09:57.
    removed_5s_rows = remove_irregular_5s_rows(
        ws,
        five["start"],
        five["end"],
        time_col=COL_TIME,
    )

    # Re-detect all blocks after row deletion because rows have shifted.
    five = block_rows(ws, "5s AVERAGED DATA", "BP DATA")

    # Safety check: stop if any irregular rows remain.
    remaining_irregular_5s_rows = find_irregular_5s_rows(
        ws,
        five["start"],
        five["end"],
        time_col=COL_TIME,
    )

    if remaining_irregular_5s_rows:
        examples = ", ".join(
            f"row {row}: {value!r}"
            for row, value in remaining_irregular_5s_rows[:5]
        )

        raise CPETAnalysisError(
            f"Irregular 5s rows were not removed correctly: {examples}"
        )

    bp = block_rows(ws, "BP DATA", "BREATH x BREATH DATA")
    bxb = block_rows(ws, "BREATH x BREATH DATA", None)

    # Clear inserted calculated-column headers/units.
    for rows in [five, bxb]:
        for c in [COL_VO2_ROLLING, COL_VO2KG_ROLLING]:
            ws.cell(rows["header"], c).value = None
            ws.cell(rows["unit"], c).value = None

    # Rolling averages.
    write_rolling_average_formulas(ws, five["start"], five["end"])
    write_rolling_average_formulas(ws, bxb["start"], bxb["end"])

    # Determine pre-exercise and peak rows.
    pre_hr_end = last_zero_load_before_exercise(
        ws,
        five["start"],
        five["end"],
        load_col=COL_LOAD,
    )

    pre_bp_row = last_zero_load_before_exercise(
        ws,
        bp["start"],
        bp["end"],
        load_col=COL_LOAD,
    )

    # These now return the LAST row where maximum load occurs.
    peak_5s_row = max_load_row(
        ws,
        five["start"],
        five["end"],
        load_col=COL_LOAD,
    )

    peak_bp_row = max_load_row(
        ws,
        bp["start"],
        bp["end"],
        load_col=COL_LOAD,
    )

    peak_bxb_row = max_load_row(
        ws,
        bxb["start"],
        bxb["end"],
        load_col=COL_LOAD,
    )

    if pre_hr_end is None:
        raise CPETAnalysisError(
            "Could not determine pre-exercise HR rows from the 5s block."
        )

    if pre_bp_row is None:
        raise CPETAnalysisError(
            "Could not determine pre-exercise BP row from the BP block."
        )

    # Final 30 sec = exactly 6 rows including the LAST peak-load row.
    last30_start = final_window_start_row(
        five["start"],
        peak_5s_row,
        window_rows=ROLLING_WINDOW_ROWS,
    )

    # Summary formulas.
    summary_rows = [
        (1, "Pre-exercise", None),
        (2, "HR", f"=AVERAGE(C{five['start']}:C{pre_hr_end})"),
        (3, "Sys", f"=E{pre_bp_row}"),
        (4, "Dia", f"=F{pre_bp_row}"),
        (5, "Peak Values", None),

        # Peak HR now includes all BxB data up to the LAST peak-load row.
        (6, "HR", f"=MAX(C{bxb['start']}:C{peak_bxb_row})"),

        (7, "Load (W)", f"=MAX(B{bxb['start']}:B{bxb['end']})"),

        # VO2 and VO2/kg rolling-average maxima up to the LAST peak-load row.
        (8, "VO2 (mL/min)", f"=MAX(G{five['start']}:G{peak_5s_row})"),
        (9, "VO2 (mL/min)/kg", f"=MAX(K{five['start']}:K{peak_5s_row})"),

        # Final 30 sec = 6 rows ending at the LAST peak-load row.
        (10, "RER", f"=AVERAGE(I{last30_start}:I{peak_5s_row})"),
        (11, "V'E", f"=AVERAGE(D{last30_start}:D{peak_5s_row})"),
        (12, "O2 pulse", f"=AVERAGE(L{last30_start}:L{peak_5s_row})"),

        # Peak BP from the LAST row where peak load occurs in BP block.
        (13, "Sys", f"=E{peak_bp_row}"),
        (14, "Dia", f"=F{peak_bp_row}"),
    ]

    for row, label, formula in summary_rows:
        ws.cell(row, 1).value = label

        if formula:
            ws.cell(row, 2).value = formula

    apply_basic_style(ws)

    ws.freeze_panes = "A16"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"Removed irregular 5s rows: {removed_5s_rows}")
    print(f"5s peak-load row used: {peak_5s_row}")
    print(f"BP peak-load row used: {peak_bp_row}")
    print(f"BxB peak-load row used: {peak_bxb_row}")
    print(f"Final 30 sec formula window: rows {last30_start}:{peak_5s_row}")
    print(f"Saved analysed workbook: {output_path}")


def default_output_path(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}_ANALYSED.xlsx")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate analysed CPET workbook from raw appended CPET file."
    )

    parser.add_argument(
        "input",
        type=Path,
        help="Raw CPET file: .xlsx, .xlsm, or tab-delimited .XLS export",
    )

    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Output .xlsx path",
    )

    args = parser.parse_args()

    input_path = args.input.expanduser().resolve()

    if not input_path.exists():
        raise FileNotFoundError(input_path)

    output_path = (
        args.output.expanduser().resolve()
        if args.output
        else default_output_path(input_path)
    )

    build_analysed_workbook(input_path, output_path)


if __name__ == "__main__":
    main()
