#!/usr/bin/env python3
"""
Generate an analysed CPET Excel workbook from a raw appended CPET file.
Includes GET, RCP detection, recovery data with formulas, and 9-panel plots.
"""

import argparse
import csv
import re
import os
import io
from datetime import datetime, time, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from cpet_logic import CPETAnalyzer
import pandas as pd

SUMMARY_ROWS = 40
ROLLING_WINDOW_ROWS = 6

COL_TIME = 1
COL_LOAD = 2
COL_HR = 3
COL_VE = 4
COL_VO2 = 6
COL_VO2_ROLLING = 7
COL_VCO2 = 8
COL_RER = 9
COL_VO2KG = 10
COL_VO2KG_ROLLING = 11
COL_O2PULSE = 12

class CPETAnalysisError(RuntimeError):
    pass

def clean_text(value: Any) -> str:
    return str(value).strip().lower() if value is not None else ""

def to_number(value: Any) -> Optional[float]:
    if value is None or value == "": return None
    if isinstance(value, (int, float)) and not isinstance(value, bool): return float(value)
    try: return float(str(value).strip())
    except: return None

def parse_time_to_seconds(value: Any) -> Optional[int]:
    if value is None: return None
    if isinstance(value, (time, datetime)): return value.hour * 3600 + value.minute * 60 + value.second
    if isinstance(value, timedelta): return int(round(value.total_seconds()))
    if isinstance(value, (int, float)):
        if 0 <= value < 1: return int(round(value * 86400))
        return None
    text = str(value).strip()
    match = re.fullmatch(r"(\d{1,2}):(\d{2})(?::(\d{2}))?", text)
    if not match: return None
    h_or_m = int(match.group(1))
    m_or_s = int(match.group(2))
    s = int(match.group(3)) if match.group(3) else None
    return h_or_m * 60 + m_or_s if s is None else h_or_m * 3600 + m_or_s * 60 + s

def coerce_cell(value: str) -> Any:
    if value is None: return None
    text = str(value).rstrip("\n\r")
    if text == "": return None
    if re.fullmatch(r"\s*\d{1,2}:(\d{2})(?::\d{2})?\s*", text): return text
    try:
        num = float(text)
        return int(num) if num.is_integer() else num
    except: return text

def load_raw_file(path: Path) -> Workbook:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}: return load_workbook(path)
    raw = path.read_text(encoding="utf-8-sig", errors="replace")
    wb = Workbook()
    ws = wb.active
    ws.title = path.stem[:31]
    reader = csv.reader(raw.splitlines(), delimiter="\t")
    for row in reader: ws.append([coerce_cell(cell) for cell in row])
    return wb

def find_row(ws, label: str) -> int:
    target = clean_text(label)
    for row in range(1, ws.max_row + 1):
        if clean_text(ws.cell(row, 1).value) == target: return row
    raise CPETAnalysisError(f"Could not find block label: {label}")

def block_rows(ws, label: str, next_label: Optional[str] = None) -> Dict[str, int]:
    label_row = find_row(ws, label)
    header_row, unit_row, data_start = label_row + 1, label_row + 2, label_row + 3
    data_end = find_row(ws, next_label) - 1 if next_label else ws.max_row
    while data_end >= data_start:
        if any(ws.cell(data_end, c).value is not None for c in range(1, 15)): break
        data_end -= 1
    return {"label": label_row, "header": header_row, "unit": unit_row, "start": data_start, "end": data_end}

def remove_irregular_5s_rows(ws, start: int, end: int) -> int:
    deleted = 0
    for row in range(end, start - 1, -1):
        secs = parse_time_to_seconds(ws.cell(row, COL_TIME).value)
        if secs is not None and secs % 5 != 0:
            ws.delete_rows(row, 1)
            deleted += 1
    return deleted

def first_load_gt_zero_row(ws, start: int, end: int) -> Optional[int]:
    for row in range(start, end + 1):
        load = to_number(ws.cell(row, COL_LOAD).value)
        if load is not None and load > 0: return row
    return None

def last_zero_load_before_exercise(ws, start: int, end: int) -> Optional[int]:
    first_ex = first_load_gt_zero_row(ws, start, end)
    search_end = first_ex - 1 if first_ex else end
    last_z = None
    for row in range(start, search_end + 1):
        if to_number(ws.cell(row, COL_LOAD).value) == 0: last_z = row
    return last_z

def zero_load_window_before_exercise(ws, start: int, end: int) -> Tuple[int, int]:
    first_ex = first_load_gt_zero_row(ws, start, end)
    search_end = first_ex - 1 if first_ex else end
    zero_rows = []
    for row in range(start, search_end + 1):
        if to_number(ws.cell(row, COL_LOAD).value) == 0: zero_rows.append(row)
    if not zero_rows: raise CPETAnalysisError("No Load=0 rows found in pre-exercise.")
    return zero_rows[0], zero_rows[-1]

def max_load_row(ws, start: int, end: int) -> int:
    m_val, m_row = None, None
    for row in range(start, end + 1):
        v = to_number(ws.cell(row, COL_LOAD).value)
        if v is not None:
            if m_val is None or v > m_val: m_val, m_row = v, row
            elif v == m_val: m_row = row
    if m_row is None: raise CPETAnalysisError("No numeric load values found.")
    return m_row

def write_rolling_average_formulas(ws, start: int, end: int) -> None:
    for row in range(start, end + 1):
        s = max(start, row - 5)
        ws.cell(row, COL_VO2_ROLLING).value = f"=AVERAGE(F{s}:F{row})"
        ws.cell(row, COL_VO2KG_ROLLING).value = f"=AVERAGE(J{s}:J{row})"
        ws.cell(row, COL_VO2_ROLLING).number_format = "0.0"
        ws.cell(row, COL_VO2KG_ROLLING).number_format = "0.0"

def apply_basic_style(ws) -> None:
    ws.column_dimensions["A"].width = 32
    for col in range(2, 16): ws.column_dimensions[get_column_letter(col)].width = 14
    bold, fill = Font(bold=True), PatternFill("solid", fgColor="D9EAF7")
    thin_gray = Side(style="thin", color="BFBFBF")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    
    for row in [1, 7, 17, 26, 35]:
        ws.cell(row, 1).font, ws.cell(row, 1).fill = bold, fill
    for row in range(1, SUMMARY_ROWS + 6):
        ws.cell(row, 1).alignment = Alignment(horizontal="left")
        ws.cell(row, 2).number_format = "0.0"
        if ws.cell(row, 1).value and "RER" in str(ws.cell(row, 1).value):
            ws.cell(row, 2).number_format = "0.00"
    
    for label in ["5s AVERAGED DATA", "BP DATA", "BREATH x BREATH DATA"]:
        try:
            r = find_row(ws, label)
            ws.cell(r, 1).font, ws.cell(r, 1).fill = bold, fill
            for c in range(1, 16):
                ws.cell(r + 1, c).font, ws.cell(r + 1, c).border = bold, border
                ws.cell(r + 2, c).font, ws.cell(r + 2, c).border = Font(italic=True), border
        except: pass

def get_df_from_ws(ws, block):
    headers = []
    for c in range(1, 26):
        val = ws.cell(block["header"], c).value
        headers.append(str(val).strip() if val is not None else f"Col{c}")
    data = []
    for r in range(block["start"], block["end"] + 1):
        data.append([ws.cell(r, c).value for c in range(1, 26)])
    return pd.DataFrame(data, columns=headers)

def build_analysed_workbook(input_path: Path, output_path: Path) -> None:
    wb = load_raw_file(input_path)
    ws = wb.active
    ws.title = f"{input_path.stem}_ANALYSED"[:31]
    ws.insert_rows(1, SUMMARY_ROWS)
    ws.insert_cols(COL_VO2_ROLLING, 1)
    ws.insert_cols(COL_VO2KG_ROLLING, 1)

    five = block_rows(ws, "5s AVERAGED DATA", "BP DATA")
    remove_irregular_5s_rows(ws, five["start"], five["end"])
    five = block_rows(ws, "5s AVERAGED DATA", "BP DATA")
    bp = block_rows(ws, "BP DATA", "BREATH x BREATH DATA")
    bxb = block_rows(ws, "BREATH x BREATH DATA", None)

    df_5s = get_df_from_ws(ws, five)
    df_bp = get_df_from_ws(ws, bp)
    df_bxb = get_df_from_ws(ws, bxb)

    analyzer = CPETAnalyzer(df_5s, df_bp, df_bxb)
    analyzer.preprocess()
    analyzer.detect_phases()
    analyzer.calculate_thresholds()
    analyzer.extract_recovery()
    
    plot_path = output_path.with_suffix(".png")
    analyzer.generate_9_panel(str(plot_path))

    write_rolling_average_formulas(ws, five["start"], five["end"])
    write_rolling_average_formulas(ws, bxb["start"], bxb["end"])

    pre_hr_end = last_zero_load_before_exercise(ws, five["start"], five["end"])
    pre_bp_row = last_zero_load_before_exercise(ws, bp["start"], bp["end"])
    rv_s, rv_e = zero_load_window_before_exercise(ws, five["start"], five["end"])
    p5, pbp, pbx = max_load_row(ws, five["start"], five["end"]), max_load_row(ws, bp["start"], bp["end"]), max_load_row(ws, bxb["start"], bxb["end"])

    def fsr(idx, block_start): return block_start + idx if idx is not None else None
    
    gsr = fsr(analyzer.results.get('GET_idx'), five['start'])
    rsr = fsr(analyzer.results.get('RCP_idx'), five['start'])
    hr1r = fsr(analyzer.results.get('HR_recovery_1min_idx'), five['start'])
    hr2r = fsr(analyzer.results.get('HR_recovery_2min_idx'), five['start'])
    bpr = fsr(analyzer.results.get('Recovery_BP_idx'), bp['start'])

    summary = [
        (1, "Pre-exercise", None), (2, "HR (bpm)", f"=AVERAGE(C{five['start']}:C{pre_hr_end})"),
        (3, "Sys (mmHg)", f"=E{pre_bp_row}"), (4, "Dia (mmHg)", f"=F{pre_bp_row}"),
        (5, "Resting VO2 (mL/kg/min)", f"=AVERAGE(K{rv_s}:K{rv_e})"),
        
        (7, "Peak Values", None), (8, "Time", f"=A{p5}"),
        (9, "HR (bpm)", f"=MAX(C{bxb['start']}:C{pbx})"), (10, "Load (W)", f"=MAX(B{bxb['start']}:B{bxb['end']})"),
        (11, "VO2 (mL/min)", f"=MAX(G{five['start']}:G{p5})"), (12, "VO2/kg", f"=MAX(K{five['start']}:K{p5})"),
        (13, "RER", f"=AVERAGE(I{max(five['start'], p5-5)}:I{p5})"), (14, "V'E", f"=AVERAGE(D{max(five['start'], p5-5)}:D{p5})"),
        (15, "SBP / DBP", f"=E{pbp}&\" / \"&F{pbp}"),

        (17, "GET (Gas Exchange Threshold)", None), (18, "Time", f"=A{gsr}" if gsr else "N/A"),
        (19, "HR (bpm)", f"=C{gsr}" if gsr else "N/A"), (20, "Load (W)", f"=B{gsr}" if gsr else "N/A"),
        (21, "VO2 (mL/min)", f"=G{gsr}" if gsr else "N/A"), (22, "VO2/kg", f"=K{gsr}" if gsr else "N/A"),
        (23, "RER", f"=I{gsr}" if gsr else "N/A"), (24, "VE/VCO2", f"=M{gsr}" if gsr else "N/A"),

        (26, "RCP (Respiratory Comp. Point)", None), (27, "Time", f"=A{rsr}" if rsr else "N/A"),
        (28, "HR (bpm)", f"=C{rsr}" if rsr else "N/A"), (29, "Load (W)", f"=B{rsr}" if rsr else "N/A"),
        (30, "VO2 (mL/min)", f"=G{rsr}" if rsr else "N/A"), (31, "VO2/kg", f"=K{rsr}" if rsr else "N/A"),
        (32, "RER", f"=I{rsr}" if rsr else "N/A"), (33, "VE/VCO2", f"=M{rsr}" if rsr else "N/A"),
        
        (35, "Recovery Data", None), 
        (36, "HR 1-min Recovery", f"=C{hr1r}" if hr1r else "N/A"),
        (37, "HR 2-min Recovery", f"=C{hr2r}" if hr2r else "N/A"), 
        (38, "End Test SBP", f"=E{bpr}" if bpr else "N/A"),
        (39, "End Test DBP", f"=F{bpr}" if bpr else "N/A"), 
        (40, "End Test HR", f"=C{bpr}" if bpr else "N/A"),
    ]
    for r, l, f in summary:
        ws.cell(r, 1).value = l
        if f is not None: ws.cell(r, 2).value = f
    
    apply_basic_style(ws)
    ws.freeze_panes = "A41"
    wb.save(output_path)

def default_output_path(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}_ANALYSED.xlsx")

def main():
    parser = argparse.ArgumentParser(description="Batch CPET Analyser v2")
    parser.add_argument("input", type=Path, nargs="+", help="Files or folders")
    parser.add_argument("-o", "--output", type=Path, help="Single output path")
    args = parser.parse_args()

    files = []
    for p in args.input:
        p = p.expanduser().resolve()
        if p.is_dir():
            files.extend([f for f in p.iterdir() if f.is_file() and f.suffix.lower() in [".xlsx", ".xls", ".xlsm"]])
        elif p.exists(): files.append(p)
    
    files = [f for f in files if "_ANALYSED" not in f.name]
    if not files: return print("No files found.")

    for i, f in enumerate(files):
        print(f"[{i+1}/{len(files)}] {f.name}...")
        out = args.output if (len(files)==1 and args.output) else default_output_path(f)
        try: build_analysed_workbook(f, out)
        except Exception as e: print(f"Error: {e}")

if __name__ == "__main__":
    main()
