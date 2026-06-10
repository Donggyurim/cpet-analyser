import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import csv
import re
import os
import io
from datetime import datetime, time, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from cpet_generate_analysed import build_analysed_workbook

class App:
    def __init__(self, root):
        self.root = root
        root.title("CPET Analyser v2")
        root.geometry("600x600")
        
        frame = ttk.Frame(root, padding="20")
        frame.pack(expand=True, fill="both")
        
        ttk.Label(frame, text="🫁 CPET Analysis Tool v2", font=("Helvetica", 18, "bold")).pack(pady=10)
        ttk.Label(frame, text="Includes Threshold Detection & 9-Panel Plots").pack()
        
        # Mode Selection
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(pady=20)
        
        ttk.Button(btn_frame, text="Single File", command=self.run_single).grid(row=0, column=0, padx=5)
        ttk.Button(btn_frame, text="Multiple Files", command=self.run_multiple).grid(row=0, column=1, padx=5)
        ttk.Button(btn_frame, text="Process Folder", command=self.run_folder).grid(row=0, column=2, padx=5)
        
        # Progress
        self.progress_var = tk.DoubleVar()
        self.progress = ttk.Progressbar(frame, variable=self.progress_var, maximum=100)
        self.progress.pack(fill="x", pady=10)
        
        # Log area
        log_frame = ttk.LabelFrame(frame, text="Activity Log")
        log_frame.pack(expand=True, fill="both", pady=10)
        
        self.log_text = tk.Text(log_frame, height=12, state="disabled", font=("Courier", 10))
        self.log_text.pack(side="left", expand=True, fill="both")
        
        scrollbar = ttk.Scrollbar(log_frame, command=self.log_text.yview)
        scrollbar.pack(side="right", fill="y")
        self.log_text.config(yscrollcommand=scrollbar.set)
        
        self.status = ttk.Label(frame, text="Ready", foreground="gray")
        self.status.pack()

    def log(self, message):
        self.log_text.config(state="normal")
        self.log_text.insert("end", f"{message}\n")
        self.log_text.config(state="disabled")
        self.log_text.see("end")
        self.root.update()

    def run_single(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls *.xlsm")])
        if file_path:
            self.process_list([Path(file_path)])

    def run_multiple(self):
        file_paths = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx *.xls *.xlsm")])
        if file_paths:
            self.process_list([Path(f) for f in file_paths])

    def run_folder(self):
        folder_path = filedialog.askdirectory()
        if folder_path:
            folder = Path(folder_path)
            files = [
                f for f in folder.iterdir() 
                if f.is_file() and f.suffix.lower() in [".xlsx", ".xls", ".xlsm"]
            ]
            files = [f for f in files if "_ANALYSED" not in f.name]
            if not files:
                messagebox.showwarning("No Files", "No Excel files found in selected folder.")
                return
            self.process_list(files)

    def process_list(self, files: List[Path]):
        total = len(files)
        self.progress_var.set(0)
        self.log(f"--- Starting Batch: {total} files ---")
        self.status.config(text="Processing...", foreground="blue")
        
        success_count = 0
        error_count = 0
        
        for i, file_path in enumerate(files):
            try:
                self.log(f"Processing: {file_path.name}...")
                output_path = file_path.with_name(f"{file_path.stem}_ANALYSED.xlsx")
                build_analysed_workbook(file_path, output_path)
                self.log(f"  ✓ Saved Excel and Plot.")
                success_count += 1
            except Exception as e:
                self.log(f"  ✗ Error: {str(e)}")
                error_count += 1
            
            self.progress_var.set(((i + 1) / total) * 100)
            self.root.update()

        self.status.config(text="Finished", foreground="green")
        self.log(f"--- Finished ---")
        self.log(f"Success: {success_count}, Errors: {error_count}")
        messagebox.showinfo("Batch Complete", f"Processed {total} files.\nSuccess: {success_count}\nErrors: {error_count}")

if __name__ == "__main__":
    root = tk.Tk()
    App(root)
    root.mainloop()
