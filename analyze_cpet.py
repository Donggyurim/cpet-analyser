import os
import sys
import csv
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def detect_file_delimiter(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            sample = f.readline() + f.readline() + f.readline()
        sniffer = csv.Sniffer()
        dialect = sniffer.sniff(sample)
        return dialect.delimiter
    except Exception:
        return '\t' if '.tsv' in file_path.lower() else ','

def find_column_by_keywords(df, keywords):
    for col in df.columns:
        col_clean = str(col).strip().lower()
        for kw in keywords:
            if kw.lower() in col_clean:
                return col
    return None

def analyze_cpet_file(input_excel_path, output_xlsx_path):
    print(f"🔄 Starting adaptive analysis on file: {input_excel_path}...")
    
    if not os.path.exists(input_excel_path):
        print(f"❌ Error: The file '{input_excel_path}' could not be found.")
        return

    # -------------------------------------------------------------
    # 1. SMART READ & STRUCTURE SEGMENTATION
    # -------------------------------------------------------------
    is_csv_format = False
    try:
        try:
            df_scan = pd.read_excel(input_excel_path, header=None, usecols=[0], engine='openpyxl')
        except Exception:
            df_scan = pd.read_excel(input_excel_path, header=None, usecols=[0], engine='xlrd')
    except Exception:
        is_csv_format = True

    lines = []
    idx_5s, idx_bp, idx_bxb = None, None, None
    
    if is_csv_format:
        chosen_sep = detect_file_delimiter(input_excel_path)
        with open(input_excel_path, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()
        for idx, line in enumerate(lines):
            if "5s AVERAGED DATA" in line: idx_5s = idx
            elif "BP DATA" in line: idx_bp = idx
            elif "BREATH x BREATH DATA" in line: idx_bxb = idx
    else:
        # If it's a real excel sheet, load column 0 elements to find strings
        df_full_excel = pd.read_excel(input_excel_path, header=None)
        for idx, row_val in enumerate(df_full_excel[0]):
            val_str = str(row_val).strip()
            if "5s AVERAGED DATA" in val_str: idx_5s = idx
            elif "BP DATA" in val_str: idx_bp = idx
            elif "BREATH x BREATH DATA" in val_str: idx_bxb = idx

    if idx_5s is None or idx_bxb is None:
        print("❌ Error: Could not find fundamental '5s AVERAGED DATA' or 'BREATH x BREATH DATA' blocks.")
        return

    # Extract tables dynamically depending on if optional BP block exists
    if is_csv_format:
        if idx_bp is not None:
            df_5s = pd.read_csv(input_excel_path, skiprows=idx_5s+1, nrows=idx_bp - idx_5s - 2, sep=chosen_sep)
            df_bp = pd.read_csv(input_excel_path, skiprows=idx_bp+1, nrows=idx_bxb - idx_bp - 2, sep=chosen_sep)
        else:
            df_5s = pd.read_csv(input_excel_path, skiprows=idx_5s+1, nrows=idx_bxb - idx_5s - 2, sep=chosen_sep)
            df_bp = pd.DataFrame() # Empty placeholder
        df_bxb = pd.read_csv(input_excel_path, skiprows=idx_bxb+1, sep=chosen_sep)
    else:
        engine_to_use = 'openpyxl' if 'openpyxl' in sys.modules else 'xlrd'
        if idx_bp is not None:
            df_5s = pd.read_excel(input_excel_path, skiprows=idx_5s+1, nrows=idx_bp - idx_5s - 2, engine=engine_to_use)
            df_bp = pd.read_excel(input_excel_path, skiprows=idx_bp+1, nrows=idx_bxb - idx_bp - 2, engine=engine_to_use)
        else:
            df_5s = pd.read_excel(input_excel_path, skiprows=idx_5s+1, nrows=idx_bxb - idx_5s - 2, engine=engine_to_use)
            df_bp = pd.DataFrame()
        df_bxb = pd.read_excel(input_excel_path, skiprows=idx_bxb+1, engine=engine_to_use)

    # -------------------------------------------------------------
    # 2. STRIP HEADERS AND DISCOVER COLUMN COORDINATES
    # -------------------------------------------------------------
    for df in [df_5s, df_bp, df_bxb]:
        if df.empty: continue
        df.dropna(how='all', axis=1, inplace=True)
        df.columns = [str(c).strip() for c in df.columns]
        df.drop(columns=[c for c in df.columns if 'Unnamed:' in c], inplace=True, errors='ignore')

    # Drop unit headers rows
    df_5s = df_5s.iloc[1:].reset_index(drop=True)
    df_bxb = df_bxb.iloc[1:].reset_index(drop=True)
    if not df_bp.empty:
        df_bp = df_bp.iloc[1:].reset_index(drop=True)

    # Map dynamic column keywords
    time_col = find_column_by_keywords(df_5s, ['Time', 'min'])
    load_col = find_column_by_keywords(df_5s, ['Load', 'W'])
    hr_col_5s = find_column_by_keywords(df_5s, ['HR', '1/min'])
    vo2_col = find_column_by_keywords(df_5s, ["V'O2", "VO2"])
    vo2_kg_col = find_column_by_keywords(df_5s, ["V'O2/kg", "VO2/kg", "(mL/min)/kg"])
    rer_col = find_column_by_keywords(df_5s, ['RER'])
    ve_col = find_column_by_keywords(df_5s, ["V'E", "VE"])
    o2_pulse_col = find_column_by_keywords(df_5s, ['O2pulse', 'O2 pulse'])
    
    hr_col_bxb = find_column_by_keywords(df_bxb, ['HR', '1/min'])

    if not load_col or not hr_col_5s or not vo2_col or not vo2_kg_col:
        print(f"❌ Error matching core metrics. Column names verified: {list(df_5s.columns)}")
        return

    # Enforce safe numeric tracking conversion
    df_5s[load_col] = pd.to_numeric(df_5s[load_col], errors='coerce').fillna(0)
    df_5s[hr_col_5s] = pd.to_numeric(df_5s[hr_col_5s], errors='coerce')
    if hr_col_bxb:
        df_bxb[hr_col_bxb] = pd.to_numeric(df_bxb[hr_col_bxb], errors='coerce')

    # Convert mapping alphabetical positions to native Excel Letter notation coordinates
    def get_excel_letter(df, col_name, offset_cols=0):
        if col_name not in df.columns: return "A"
        idx = list(df.columns).index(col_name) + 1 + offset_cols
        return get_column_letter(idx)

    # Calculate target letter coordinates based completely on current file layout
    load_let = get_excel_letter(df_5s, load_col)
    hr_let_bxb = get_excel_letter(df_bxb, hr_col_bxb)
    
    # We will inject rolling average variables in newly created open spaces directly adjacent to targets
    vo2_let = get_excel_letter(df_5s, vo2_col)
    vo2_kg_let = get_excel_letter(df_5s, vo2_kg_col)
    
    # Find positions of other columns to reference in averages
    rer_let = get_excel_letter(df_5s, rer_col)
    ve_let = get_excel_letter(df_5s, ve_col)
    o2_pulse_let = get_excel_letter(df_5s, o2_pulse_col)

    # -------------------------------------------------------------
    # 3. RUN CALCULATIONS
    # -------------------------------------------------------------
    peak_load_value = df_5s[load_col].max()
    idx_peak_5s = df_5s[df_5s[load_col] == peak_load_value].index[0]

    df_pre = df_5s[df_5s[time_col].astype(str).str.strip() <= '02:00']
    pre_hr_val = df_pre[hr_col_5s].mean()

    # Track BP parameters dynamically if segment structural features are available
    pre_sys_val, pre_dia_val, peak_sys_val, peak_dia_val = "N/A", "N/A", "N/A", "N/A"
    if not df_bp.empty:
        load_col_bp = find_column_by_keywords(df_bp, ['Load', 'W'])
        time_col_bp = find_column_by_keywords(df_bp, ['Time', 'min'])
        sys_col_bp = find_column_by_keywords(df_bp, ['Sys', 'Psys', 'mmHg'])
        dia_col_bp = find_column_by_keywords(df_bp, ['Dia', 'Pdia', 'mmHg'])
        
        if load_col_bp and sys_col_bp and dia_col_bp:
            df_bp[load_col_bp] = pd.to_numeric(df_bp[load_col_bp], errors='coerce').fillna(0)
            df_bp[sys_col_bp] = pd.to_numeric(df_bp[sys_col_bp], errors='coerce')
            df_bp[dia_col_bp] = pd.to_numeric(df_bp[dia_col_bp], errors='coerce')
            
            df_pre_bp = df_bp[df_bp[load_col_bp] == 0]
            df_pre_bp = df_pre_bp[df_pre_bp[time_col_bp].astype(str).str.strip() <= '02:30']
            pre_sys_val = df_pre_bp[sys_col_bp].dropna().iloc[-1] if not df_pre_bp[sys_col_bp].dropna().empty else "N/A"
            pre_dia_val = df_pre_bp[dia_col_bp].dropna().iloc[-1] if not df_pre_bp[dia_col_bp].dropna().empty else "N/A"
            
            bp_peak_row = df_bp[df_bp[load_col_bp] == peak_load_value]
            if not bp_peak_row.empty:
                peak_sys_val = bp_peak_row[sys_col_bp].dropna().iloc[0] if not bp_peak_row[sys_col_bp].dropna().empty else "N/A"
                peak_dia_val = bp_peak_row[dia_col_bp].dropna().iloc[0] if not bp_peak_row[dia_col_bp].dropna().empty else "N/A"
            else:
                max_bp_load_idx = df_bp[load_col_bp].idxmax()
                peak_sys_val = df_bp.loc[max_bp_load_idx, sys_col_bp]
                peak_dia_val = df_bp.loc[max_bp_load_idx, dia_col_bp]

    # -------------------------------------------------------------
    # 4. SHEET RE-CONSTRUCTION ENGINE
    # -------------------------------------------------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CPET_ANALYSED"

    # Base Summary blocks layout
    ws['A1'] = "Pre-exercise"
    ws['A1'].font = Font(bold=True)
    ws['A2'] = "HR"
    ws['B2'] = pre_hr_val
    ws['A3'] = "Sys"
    ws['B3'] = pre_sys_val
    ws['A4'] = "Dia"
    ws['B4'] = pre_dia_val

    ws['A5'] = "Peak Values"
    ws['A5'].font = Font(bold=True)
    
    offset = 17 
    start_row_5s = offset + 2  
    peak_row_excel = start_row_5s + idx_peak_5s
    
    total_5s_rows = (idx_bp if idx_bp is not None else idx_bxb) - idx_5s
    start_row_bxb = offset + total_5s_rows + 2

    # Map dynamic column letters directly inside formulas string coordinates
    ws['A6'] = "HR"
    ws['B6'] = f"=MAX({hr_let_bxb}{start_row_bxb}:{hr_let_bxb}{start_row_bxb + len(df_bxb)})" 
    ws['A7'] = "Load (W)"
    ws['B7'] = f"=MAX({load_let}{start_row_5s}:{load_let}{start_row_5s + len(df_5s) - 1})"
    
    # We will compute rolling averages inside helper columns custom positioned at the end of the sheet
    rolling_vo2_let = get_column_letter(len(df_5s.columns) + 2)
    rolling_vo2_kg_let = get_column_letter(len(df_5s.columns) + 3)

    ws['A8'] = "VO2 (mL/min)"
    ws['B8'] = f"=MAX({rolling_vo2_let}{start_row_5s}:{rolling_vo2_let}{start_row_5s + len(df_5s) - 1})" 
    ws['A9'] = "VO2 (mL/min)/kg"
    ws['B9'] = f"=MAX({rolling_vo2_kg_let}{start_row_5s}:{rolling_vo2_kg_let}{start_row_5s + len(df_5s) - 1})" 
    
    ws['A10'] = "RER"
    ws['B10'] = f"=AVERAGE({rer_let}{peak_row_excel - 5}:{rer_let}{peak_row_excel})"
    ws['A11'] = "V'E"
    ws['B11'] = f"=AVERAGE({ve_let}{peak_row_excel - 5}:{ve_let}{peak_row_excel})"
    ws['A12'] = "O2 pulse"
    ws['B12'] = f"=AVERAGE({o2_pulse_let}{peak_row_excel - 5}:{o2_pulse_let}{peak_row_excel})"
    
    ws['A13'] = "Sys"
    ws['B13'] = peak_sys_val
    ws['A14'] = "Dia"
    ws['B14'] = peak_dia_val

    # -------------------------------------------------------------
    # 5. STREAM STRUCTURE ARRAYS SAFELY BACK
    # -------------------------------------------------------------
    if is_csv_format:
        for r_idx, line in enumerate(lines[idx_5s:]):
            curr_row = offset + r_idx
            row_values = line.strip().split(chosen_sep)
            for c_idx, val in enumerate(row_values):
                val_clean = val.strip()
                if val_clean:
                    try: ws.cell(row=curr_row, column=c_idx+1, value=float(val_clean))
                    except ValueError: ws.cell(row=curr_row, column=c_idx+1, value=val_clean)
    else:
        df_full_raw = pd.read_excel(input_excel_path, header=None, engine=engine_to_use)
        for r_idx, row in df_full_raw.iloc[idx_5s:].iterrows():
            curr_row = offset + (r_idx - idx_5s)
            for c_idx, cell_value in enumerate(row):
                if pd.notna(cell_value): ws.cell(row=curr_row, column=c_idx+1, value=cell_value)

    # Set up rolling headings out to the right side safely away from overlapping
    ws.cell(row=offset+1, column=len(df_5s.columns) + 2, value="VO2_Rolling30s")
    ws.cell(row=offset+1, column=len(df_5s.columns) + 3, value="VO2/kg_Rolling30s")

    for r in range(start_row_5s, start_row_5s + len(df_5s)):
        back_steps = 5 if (r >= start_row_5s + 5) else (r - start_row_5s)
        ws[f"{rolling_vo2_let}{r}"] = f"=AVERAGE({vo2_let}{r - back_steps}:{vo2_let}{r})"
        ws[f"{rolling_vo2_kg_let}{r}"] = f"=AVERAGE({vo2_kg_let}{r - back_steps}:{vo2_kg_let}{r})"

    ws.column_dimensions['A'].width = 24
    wb.save(output_xlsx_path)
    print(f"🎉 Success! Generated perfect adaptive matrix: {output_xlsx_path}\n")

import sys
input_file = sys.argv[1]
output_file = sys.argv[2]
analyze_cpet_file(input_file, output_file)
