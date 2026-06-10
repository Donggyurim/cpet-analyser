import streamlit as st
import pandas as pd
import io
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Import our core logic
from cpet_generate_analysed import build_analysed_workbook

st.set_page_config(page_title="CPET Analyser v2", page_icon="🫁", layout="wide")

st.title("🫁 CPET Analysis Tool v2")
st.markdown("""
### Advanced CPET Analysis
- **9-Panel Wasserman Plots**: Visualise Gas Exchange Threshold (GET) and Respiratory Compensation Point (RCP).
- **Automated Thresholds**: GET detection via V-slope; RCP detection via ventilatory equivalents.
- **Recovery Analysis**: HR recovery at 1 & 2 mins, and end-of-test blood pressure.
- **Batch Processing**: Upload multiple files to get a ZIP of all reports and plots.
""")

uploaded_files = st.file_uploader("Upload raw CPET file(s)", type=["xlsx", "xlsm", "xls"], accept_multiple_files=True)

if uploaded_files:
    if st.button("🚀 Process All Files"):
        results = []
        progress_bar = st.progress(0)
        status = st.empty()
        
        # Temp directory for processing
        temp_dir = Path("temp_analysis")
        temp_dir.mkdir(exist_ok=True)

        try:
            for i, uploaded_file in enumerate(uploaded_files):
                status.text(f"Processing {i+1}/{len(uploaded_files)}: {uploaded_file.name}...")
                
                # Save uploaded file to temp
                input_path = temp_dir / uploaded_file.name
                with open(input_path, "wb") as f:
                    f.write(uploaded_file.getbuffer())
                
                output_path = temp_dir / f"{input_path.stem}_ANALYSED.xlsx"
                plot_path = temp_dir / f"{input_path.stem}_ANALYSED.png"
                
                # Process
                build_analysed_workbook(input_path, output_path)
                
                # Read results back
                with open(output_path, "rb") as f:
                    xlsx_data = f.read()
                with open(plot_path, "rb") as f:
                    png_data = f.read()
                    
                results.append({
                    "name": uploaded_file.name,
                    "xlsx": xlsx_data,
                    "png": png_data,
                    "xlsx_name": output_path.name,
                    "png_name": plot_path.name
                })
                
                progress_bar.progress((i + 1) / len(uploaded_files))

            status.text("Done!")
            
            if len(results) == 1:
                res = results[0]
                col1, col2 = st.columns(2)
                with col1:
                    st.success(f"Analysis complete for {res['name']}")
                    st.download_button("📥 Download Excel Report", res['xlsx'], res['xlsx_name'])
                    st.download_button("🖼️ Download 9-Panel Plot", res['png'], res['png_name'])
                with col2:
                    st.image(res['png'], caption="9-Panel Wasserman Plot")
            else:
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, "w") as zf:
                    for res in results:
                        zf.writestr(res['xlsx_name'], res['xlsx'])
                        zf.writestr(res['png_name'], res['png'])
                
                st.success(f"Processed {len(results)} files successfully!")
                st.download_button("📥 Download All Results (ZIP)", zip_buffer.getvalue(), "CPET_Batch_Results.zip")

        except Exception as e:
            st.error(f"Error: {e}")
            st.exception(e)
        finally:
            # Cleanup temp
            for f in temp_dir.glob("*"): f.unlink()
            temp_dir.rmdir()
